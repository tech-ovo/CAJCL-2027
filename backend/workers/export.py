"""Export the database. Runs on Modal, and unmodified in a Google Colab.

THIS IS THE FALLBACK EVERY OTHER FALLBACK RESTS ON. Given a .db file and a
couple of arguments it must run anywhere, with no repository checkout, no
Modal, and no network. That is a hard architectural rule -- see docs/stack.md.

    In a Colab:
        !pip install openpyxl
        # upload cajcl.db, then:
        !python export.py --db cajcl.db --out ./exports

FOUR FILES PER EXPORT
    Excel and SQL, each in a FULL version with personal information and an
    ANONYMISED version showing only user IDs. The anonymised versions exist so
    they can be handed to an AI or an outside helper without exposing minors'
    data -- which is a thing that will be wanted at 2am during convention, and
    the wrong moment to be inventing a redaction scheme.

WHAT ANONYMISED MEANS HERE -- PRECISELY
    Every ATTENDEE's personal information is dropped entirely: names,
    guardians, emails, phones, free-text notes, and the raw pasted roster text.
    Dropped, not masked and not hashed into something reversible. What remains
    is ids, chapters, levels, selections and counts -- everything anyone doing
    analysis actually needs.

    It does NOT strip the convention's own configuration. `settings` still
    carries the remit-to line, which names the treasurer, and `documents` still
    carries the printed prose. Both are already printed on every invoice that
    goes in the post, so they are public by construction -- but be aware that
    "anonymised" here means "no student, parent, or volunteer data", not "no
    proper nouns at all". A test asserts the first of those; nothing can assert
    the second, because a chapter's own name is data the export exists to carry.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import pathlib
import sqlite3
import sys

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

# Columns removed from the anonymised exports, by table.
REDACT = {
    "people": ["first_name", "middle_name", "last_name", "suffix", "raw_name_input",
               "email", "cell_phone", "guardian_name", "guardian_phone",
               "availability_note", "code_hmac"],
    # discount_reason is free text an admin types and could name a family.
    "schools": ["drive_folder_id", "notes", "discount_reason"],
    "audit_log": ["summary", "value_detail", "ip_hash"],
    "roster_imports": ["raw_text"],
    "sessions": ["token_hash", "user_agent", "ip_hash"],
    "login_attempts": ["attempted_code_hmac", "ip_hash"],
    "payments": ["note", "reference"],
    "contest_submissions": ["original_name"],
}

# Never exported at all, in either version. A session token is a live
# credential and an export is a file that gets emailed around.
SKIP_TABLES = {"sessions", "login_attempts", "schema_migrations"}

# Prose is redacted from the anonymised export as well, because a person's name
# reaches these tables by a route the column-level rules above cannot see:
# `invoice.remit_to` names the treasurer, and the treasurer is usually also a
# sponsor on somebody's roster. The NUMBERS survive -- fees, ratios, deadlines,
# flags -- which is the part an analysis actually needs.
PROSE_VALUE_TYPES = {"string", "markdown"}


def open_db(path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def table_names(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table' "
        "AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
    return [r["name"] for r in rows if r["name"] not in SKIP_TABLES]


def columns_of(conn: sqlite3.Connection, table: str) -> list[str]:
    return [r["name"] for r in conn.execute(f'PRAGMA table_info("{table}")')]


def kept_columns(conn: sqlite3.Connection, table: str, anonymized: bool) -> list[str]:
    """Always EVERY column, in both versions.

    The anonymised export blanks values rather than dropping columns, because
    dropping them produces a dump that will not restore: `audit_log.summary` is
    NOT NULL, so an INSERT omitting it fails, and the file that exists precisely
    to be handed to an outside helper cannot be loaded by one. Same schema,
    same shape, no personal data.
    """
    return columns_of(conn, table)


def not_null_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {r["name"] for r in conn.execute(f'PRAGMA table_info("{table}")')
            if r["notnull"]}


def unique_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    """Columns covered by a UNIQUE index.

    A redacted value in one of these has to stay distinct per row, or the
    restore fails on the first duplicate -- `people.code_hmac` is uniquely
    indexed, so blanking every row to the same placeholder makes the dump
    unloadable after exactly one person.
    """
    out: set[str] = set()
    for index in conn.execute(f'PRAGMA index_list("{table}")'):
        if not index["unique"]:
            continue
        for column in conn.execute(f'PRAGMA index_info("{index["name"]}")'):
            if column["name"]:
                out.add(column["name"])
    return out


def rows_of(conn: sqlite3.Connection, table: str, keep: list[str],
            anonymized: bool):
    """Every row of a table, redacted in place for the anonymised export.

    A redacted value becomes NULL where the column allows it and the string
    `[redacted]` where it does not, so the row still satisfies every constraint
    the schema declares.
    """
    quoted = ", ".join(f'"{c}"' for c in keep)
    rows = conn.execute(f'SELECT {quoted} FROM "{table}"').fetchall()
    if not anonymized:
        for record in rows:
            yield {c: record[c] for c in keep}
        return

    required = not_null_columns(conn, table)
    unique = unique_columns(conn, table)
    redact = set(REDACT.get(table, []))

    for position, record in enumerate(rows, start=1):
        values = {c: record[c] for c in keep}
        for column in redact:
            if column not in values or values[column] is None:
                continue
            if column in unique:
                values[column] = f"[redacted-{position}]"
            elif column in required:
                values[column] = "[redacted]"
            else:
                values[column] = None
        if table == "settings" and values.get("value_type") in PROSE_VALUE_TYPES:
            values["value"] = "[redacted]"
        elif table in ("documents", "announcements"):
            values["body_md"] = "[redacted]"
        yield values


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------

def export_sql(conn: sqlite3.Connection, out: pathlib.Path, anonymized: bool) -> pathlib.Path:
    """A real SQLite dump that opens in DB Browser and restores with `sqlite3 <`.

    Written by hand rather than with iterdump() because the anonymised version
    has to redact values, which iterdump cannot do.
    """
    path = out / f"cajcl-{stamp()}{'-anonymized' if anonymized else '-full'}.sql"
    with path.open("w", encoding="utf-8") as handle:
        handle.write("-- CAJCL 2027 export\n")
        handle.write(f"-- {dt.datetime.now(dt.timezone.utc).isoformat()}\n")
        handle.write(f"-- {'ANONYMISED' if anonymized else 'FULL - contains personal information'}\n")
        handle.write("PRAGMA foreign_keys = OFF;\nBEGIN TRANSACTION;\n\n")

        for row in conn.execute(
                "SELECT sql FROM sqlite_master WHERE type IN ('table','index','trigger') "
                "AND sql IS NOT NULL AND name NOT LIKE 'sqlite_%'"):
            handle.write(row["sql"].rstrip().rstrip(";") + ";\n")
        handle.write("\n")

        for table in table_names(conn):
            keep = kept_columns(conn, table, anonymized)
            quoted = ", ".join(f'"{c}"' for c in keep)
            for record in rows_of(conn, table, keep, anonymized):
                values = ", ".join(sql_literal(record[c]) for c in keep)
                handle.write(f'INSERT INTO "{table}" ({quoted}) VALUES ({values});\n')

        handle.write("\nCOMMIT;\n")
    return path


def sql_literal(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, bytes):
        return "X'" + value.hex() + "'"
    return "'" + str(value).replace("'", "''") + "'"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_xlsx(conn: sqlite3.Connection, out: pathlib.Path, anonymized: bool) -> pathlib.Path:
    from openpyxl import Workbook

    path = out / f"cajcl-{stamp()}{'-anonymized' if anonymized else '-full'}.xlsx"
    book = Workbook()
    book.remove(book.active)

    # A readable summary first, because whoever opens this in a hurry wants the
    # numbers before the tables.
    sheet = book.create_sheet("Summary")
    sheet.append(["CAJCL 2027 export"])
    sheet.append(["Generated", dt.datetime.now(dt.timezone.utc).isoformat()])
    sheet.append(["Contents", "Anonymised" if anonymized else
                              "FULL - contains personal information"])
    sheet.append([])
    sheet.append(["Chapter", "Level", "Delegates", "Adults", "Complete",
                  "Owed", "Paid", "Balance"])
    for row in conn.execute(
            "SELECT s.name, s.level, ss.delegates_active, ss.adults_active, "
            "       ss.delegates_complete, ss.amount_owed_cents, ss.amount_paid_cents "
            "FROM schools s JOIN school_stats ss ON ss.school_id = s.id "
            "WHERE s.kind = 'chapter' ORDER BY s.name"):
        owed = row["amount_owed_cents"] or 0
        paid = row["amount_paid_cents"] or 0
        sheet.append([row["name"], row["level"], row["delegates_active"],
                      row["adults_active"], row["delegates_complete"],
                      owed / 100, paid / 100, (owed - paid) / 100])

    for table in table_names(conn):
        keep = kept_columns(conn, table, anonymized)
        # Excel sheet names cap at 31 characters.
        tab = book.create_sheet(table[:31])
        tab.append(keep)
        for record in rows_of(conn, table, keep, anonymized):
            tab.append([record[c] for c in keep])

    book.save(path)
    return path


def stamp() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d-%H%M")


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def run(fmt: str = "xlsx", anonymized: bool = False, db_path: str | None = None,
        out_dir: str = "/tmp/exports") -> dict:
    """Called by Modal. Returns a description of what was written."""
    path = db_path or os.environ.get("EXPORT_DB_PATH")
    if path is None:
        # On Modal the database is Turso, so mirror it to a local file first.
        # Everything downstream then works on a real SQLite file, which is what
        # makes the Colab fallback identical to production.
        path = mirror_turso_to_file(out_dir)

    out = pathlib.Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    conn = open_db(path)
    try:
        written = (export_sql(conn, out, anonymized) if fmt == "sql"
                   else export_xlsx(conn, out, anonymized))
    finally:
        conn.close()

    return {"path": str(written), "bytes": written.stat().st_size,
            "format": fmt, "anonymized": anonymized}


def mirror_turso_to_file(out_dir: str) -> str:
    """Copy the live database into a local SQLite file.

    libSQL *is* SQLite, so this produces a real .db that opens in DB Browser and
    loads into a Colab -- which is the property every fallback plan depends on.
    """
    import os

    import libsql

    url = os.environ["TURSO_DATABASE_URL"]
    token = os.environ.get("TURSO_AUTH_TOKEN")
    client = libsql.connect(url, auth_token=token or "", isolation_level=None)

    pathlib.Path(out_dir).mkdir(parents=True, exist_ok=True)
    local = str(pathlib.Path(out_dir) / "mirror.db")
    if pathlib.Path(local).exists():
        pathlib.Path(local).unlink()

    target = sqlite3.connect(local)
    try:
        for (statement,) in client.execute(
                "SELECT sql FROM sqlite_master WHERE sql IS NOT NULL "
                "AND name NOT LIKE 'sqlite_%'").fetchall():
            target.execute(statement)

        tables = [row[0] for row in client.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%'").fetchall()]

        for table in tables:
            cursor = client.execute(f'SELECT * FROM "{table}"')
            rows = cursor.fetchall()
            if not rows:
                continue
            columns = [column[0] for column in cursor.description]
            marks = ", ".join("?" * len(columns))
            names = ", ".join(f'"{c}"' for c in columns)
            target.executemany(
                f'INSERT INTO "{table}" ({names}) VALUES ({marks})',
                [tuple(row) for row in rows])
        target.commit()
    finally:
        target.close()
        client.close()
    return local


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", default=None,
                    help="path to a SQLite .db file "
                         "(or set EXPORT_DB_PATH, which run() already reads)")
    ap.add_argument("--out", default="./exports")
    ap.add_argument("--only", choices=["sql", "xlsx"], default=None,
                    help="write just one format instead of all four files")
    args = ap.parse_args()

    # Same fallback `run()` uses, so the documented environment variable means
    # the same thing whether this file is imported or run from a terminal.
    db = args.db or os.environ.get("EXPORT_DB_PATH")
    if db is None:
        ap.error("give --db, or set EXPORT_DB_PATH")

    out = pathlib.Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    conn = open_db(db)

    formats = [args.only] if args.only else ["sql", "xlsx"]
    try:
        for fmt in formats:
            for anonymized in (False, True):
                writer = export_sql if fmt == "sql" else export_xlsx
                path = writer(conn, out, anonymized)
                print(f"  {path.name}  ({path.stat().st_size / 1024:.0f} KB)")
    finally:
        conn.close()

    print(f"\nwritten to {out.resolve()}")
    print("The -full files contain names and guardian contact details.")
    print("The -anonymized files do not, and are the ones safe to share.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
